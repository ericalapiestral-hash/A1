# -*- coding: utf-8 -*-
# 12V 펌프 2개 버전 — 전체 결선표 엑셀 생성기
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '12V_펌프2개_전체결선표.xlsx')

# ---- 스타일 ----
FONT_BASE = Font(name='맑은 고딕', size=11)
FONT_BOLD = Font(name='맑은 고딕', size=11, bold=True)
FONT_TITLE = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
FONT_H2 = Font(name='맑은 고딕', size=12, bold=True, color='0F4C81')
FONT_WARN = Font(name='맑은 고딕', size=11, bold=True, color='B0291F')
FONT_OK = Font(name='맑은 고딕', size=11, bold=True, color='1F6B43')

FILL_TITLE = PatternFill('solid', start_color='0F4C81')
FILL_HDR_A = PatternFill('solid', start_color='DCEBFA')
FILL_HDR_B = PatternFill('solid', start_color='FDE6E0')
FILL_HDR_C = PatternFill('solid', start_color='E2F0E4')
FILL_HDR_D = PatternFill('solid', start_color='FFF4D6')
FILL_ALT   = PatternFill('solid', start_color='F7F7F7')
FILL_NOTE  = PatternFill('solid', start_color='FFF8E6')
FILL_DANGER= PatternFill('solid', start_color='FFE5E0')

THIN = Side(style='thin', color='B0B7C0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, text, span, fill=FILL_TITLE):
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = CENTER
    ws.row_dimensions[row].height = 26
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)

def section_row(ws, row, text, span, fill=FILL_HDR_A):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_H2; c.fill = fill; c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)

def header_row(ws, row, headers, fill=FILL_HDR_A):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_BOLD; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 20

def data_rows(ws, start_row, rows, alt=True):
    r = start_row
    for idx, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = FONT_BASE
            c.alignment = CENTER if j <= 2 else LEFT
            c.border = BORDER
            if alt and idx % 2 == 1:
                c.fill = FILL_ALT
        ws.row_dimensions[r].height = 22
        r += 1
    return r

def note_block(ws, row, text, span, fill=FILL_NOTE, font=None):
    c = ws.cell(row=row, column=1, value=text)
    c.font = font or FONT_BASE; c.fill = fill; c.alignment = LEFT; c.border = BORDER
    ws.row_dimensions[row].height = max(22, 16 * (1 + text.count('\n')))
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)

wb = Workbook()

# ============ 시트 1: 개요 ============
ws = wb.active
ws.title = '개요'
set_widths(ws, [22, 60])
title_row(ws, 1, '12V 펌프 2개 버전 — 전체 결선표', 2)
r = 3
ws.cell(r,1,'구성').font = FONT_BOLD
ws.cell(r,2,'A보드(센서·LCD) + B보드(12V 펌프1·12V 펌프2·220V 히터·BLE)').font = FONT_BASE
r += 1
ws.cell(r,1,'통신').font = FONT_BOLD
ws.cell(r,2,'보드 간 ESP-NOW(채널1, 무선) · 폰 ↔ B보드 BLE(무선) — 전선 없음').font = FONT_BASE
r += 1
ws.cell(r,1,'대상 코드').font = FONT_BOLD
ws.cell(r,2,'A_board/A_board.ino  ·  B_board/B_board.ino').font = FONT_BASE
r += 2
section_row(ws, r, '시트 구성', 2); r += 1
sheet_list = [
    ('A보드_결선',   '센서 5개 + LCD + 전원 레일 (브레드보드)'),
    ('B보드_결선',   '릴레이 + 12V 펌프 2개(+다이오드) + 220V 히터'),
    ('핀_요약',     'A보드/B보드 ESP32 핀 사용 한눈에'),
    ('부품_목록',   '12V 펌프 2개 버전 부품 리스트'),
    ('주의사항',    '자주 헷갈리는 7가지 + 안전'),
]
header_row(ws, r, ['시트', '내용']); r += 1
r = data_rows(ws, r, sheet_list)

r += 1
note_block(ws, r,
  '* 무선 통신(ESP-NOW, BLE)은 전선이 없으므로 결선표에 없음.\n'
  '* 양 보드 GND는 보드별로 1점 공통 (보드끼리는 무선이라 분리 OK).\n'
  '* 220V 작업은 반드시 콘센트(돼지코) 뽑은 상태에서.',
  2)

# ============ 시트 2: A보드 ============
ws = wb.create_sheet('A보드_결선')
set_widths(ws, [22, 28, 28, 40])
title_row(ws, 1, 'A보드 — ESP32 38핀 DevKit + 확장보드 + 브레드보드', 4)

r = 3
section_row(ws, r, '① 전원 레일 (브레드보드에 먼저 깔기)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['항목', '출발', '도착', '비고'], FILL_HDR_A); r += 1
rows = [
 ('5V 공급',  'ESP32 5V (VIN)', '브레드보드 5V 레일', '5V 센서·LCD 공급용'),
 ('3.3V 공급','ESP32 3V3',      '브레드보드 3.3V 레일', '수온·조도 (native 3.3V)'),
 ('GND 공급', 'ESP32 GND',      '브레드보드 GND 레일', '모든 GND 공통 1점 접지'),
 ('디커플링1', '5V 레일',         'GND 레일',  '10µF 전해 + 0.1µF 세라믹 (병렬, ESP32 근처)'),
 ('디커플링2', '3.3V 레일',       'GND 레일',  '10µF 전해 + 0.1µF 세라믹 (병렬)'),
]
r = data_rows(ws, r, rows); r += 1

section_row(ws, r, '② 수온 — DS18B20 (3.3V, OneWire)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['센서 핀', '도착', 'ESP32 핀', '비고'], FILL_HDR_A); r += 1
rows = [
 ('VCC (빨강)', '3.3V 레일', '—', ''),
 ('GND (검정)', 'GND 레일',  '—', ''),
 ('DATA (노랑)','—',         'IO4', '4.7kΩ 풀업 필수: DATA ↔ 3.3V 레일 (없으면 −127℃)'),
]
r = data_rows(ws, r, rows); r += 1

section_row(ws, r, '③ 조도 — GL5528 (3.3V, 분압 아날로그)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['부품/핀', '한쪽', '반대쪽', '비고'], FILL_HDR_A); r += 1
rows = [
 ('GL5528 다리1', '3.3V 레일', '—', ''),
 ('GL5528 다리2', 'IO33 (분압 노드)', '—', '3.3V–LDR–IO33–10k–GND 분압'),
 ('10kΩ 저항',    'IO33',     'GND 레일', '분압 하단 저항'),
 ('0.1µF 세라믹', 'IO33 (노드)', 'GND 레일', '필터, ADC 안정'),
]
r = data_rows(ws, r, rows); r += 1

section_row(ws, r, '④ 탁도 — SEN0189 (5V, 분압)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['부품/핀', '한쪽', '반대쪽', '비고'], FILL_HDR_A); r += 1
rows = [
 ('센서 VCC', '5V 레일', '—', ''),
 ('센서 GND', 'GND 레일','—', ''),
 ('센서 OUT', '10kΩ 저항 한쪽', '—', ''),
 ('10kΩ 저항', '센서 OUT', 'IO32 (분압 노드)', '직렬 저항'),
 ('20kΩ 저항', 'IO32', 'GND 레일', '하단 저항'),
 ('0.1µF 세라믹', 'IO32 (노드)', 'GND 레일', '필터'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* 분압 ×0.67 → 코드 DIV_TURB = 1.5 로 복원', 4); r += 2

section_row(ws, r, '⑤ pH — PH4502C (5V, 분압, 입력전용 IO34)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['부품/핀', '한쪽', '반대쪽', '비고'], FILL_HDR_A); r += 1
rows = [
 ('센서 V+', '5V 레일', '—', ''),
 ('센서 GND', 'GND 레일', '—', ''),
 ('센서 PO', '10kΩ 저항 한쪽', '—', '아날로그 출력'),
 ('10kΩ 저항', '센서 PO', 'IO34 (분압 노드)', '직렬 저항'),
 ('20kΩ 저항', 'IO34', 'GND 레일', '하단 저항'),
 ('1µF (전해/세라믹)', 'IO34 (노드)', 'GND 레일', 'pH 안정용 (0.1µF보다 큼)'),
 ('To, 온도핀', '—', '—', '미사용 (연결 X)'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* 분압 ×0.67 → 코드 DIV_PH = 1.5 로 복원', 4); r += 2

section_row(ws, r, '⑥ 수위 — XKC-Y25-V (5V, 디지털 읽기)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['센서 선', '한쪽', '반대쪽', '비고'], FILL_HDR_A); r += 1
rows = [
 ('갈색 (VCC)', '5V 레일', '—', ''),
 ('파랑 (GND)', 'GND 레일','—', ''),
 ('검정 (OUT)', '10kΩ 저항 한쪽', '—', ''),
 ('10kΩ 저항',  '센서 OUT', 'IO35 (분압 노드)', '5V 로직 → 3.3V 안전 강하'),
 ('20kΩ 저항',  'IO35', 'GND 레일', ''),
 ('노랑 (모드)', '—', '—', '미사용 (연결 X)'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* digitalRead로 읽음: 물 차면 1(정상), 부족 0', 4); r += 2

section_row(ws, r, '⑦ LCD 16×2 (HD44780, 5V, 4비트 모드)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['LCD 핀#', '이름', '도착', '비고'], FILL_HDR_A); r += 1
rows = [
 (1,'VSS','GND 레일',''),
 (2,'VDD','5V 레일',''),
 (3,'V0','10kΩ 가변저항 가운데(와이퍼)','글자 진하기 조절'),
 (4,'RS','ESP32 IO13',''),
 (5,'RW','GND 레일','항상 쓰기 모드'),
 (6,'E','ESP32 IO14',''),
 ('7~10','D0~D3','미사용','4비트 모드'),
 (11,'D4','ESP32 IO27',''),
 (12,'D5','ESP32 IO16',''),
 (13,'D6','ESP32 IO17',''),
 (14,'D7','ESP32 IO18',''),
 (15,'A (백라이트+)','5V 레일','220Ω 직렬 권장'),
 (16,'K (백라이트−)','GND 레일',''),
]
r = data_rows(ws, r, rows); r += 1
header_row(ws, r, ['가변저항 단자', '도착', '', ''], FILL_HDR_A); r += 1
rows = [
 ('양끝 단자 (둘 중 하나)', '5V 레일', '', ''),
 ('양끝 단자 (나머지)',     'GND 레일','', ''),
 ('가운데 (와이퍼)',         'LCD 3번 핀 (V0)', '', ''),
]
r = data_rows(ws, r, rows); r += 1

# ============ 시트 3: B보드 ============
ws = wb.create_sheet('B보드_결선')
set_widths(ws, [22, 32, 32, 40])
title_row(ws, 1, 'B보드 — ESP32 WROOM-32 + 3채널 릴레이 모듈 (LOW=ON)', 4)

r = 3
section_row(ws, r, '① 전원 / 릴레이 모듈', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['항목', '출발', '도착', '비고'], FILL_HDR_B); r += 1
rows = [
 ('릴레이 VCC', 'ESP32 5V', '릴레이 모듈 VCC', '릴레이 코일 구동'),
 ('릴레이 GND', 'ESP32 GND','릴레이 모듈 GND', '공통 GND'),
 ('디커플링',   '릴레이 VCC', '릴레이 GND', '10µF + 0.1µF 병렬 (릴레이에 최대한 가까이)'),
 ('ESP32 전원', 'USB 5V 어댑터 (1.5~2A)', 'ESP32 USB-C', '상시 급전'),
]
r = data_rows(ws, r, rows); r += 1

section_row(ws, r, '② ESP32 → 릴레이 제어선', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['ESP32 핀', '릴레이 IN', '채널', '용도 / 트리거'], FILL_HDR_B); r += 1
rows = [
 ('IO25', 'IN1', 'CH1', '펌프1 (급수) · LOW=ON'),
 ('IO26', 'IN2', 'CH2', '펌프2 (순환) · LOW=ON'),
 ('IO27', 'IN3', 'CH3', '히터 (220V) · LOW=ON'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* 릴레이가 반대로 동작하면 코드의 RELAY_LOW_TRIG = false 로 변경', 4); r += 2

section_row(ws, r, '③ 펌프1 — 급수 (12V DC · 릴레이 CH1) ★ 다이오드 포함', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['항목', '출발', '도착', '비고'], FILL_HDR_B); r += 1
rows = [
 ('12V 활선 입력', '12V 어댑터 +', '릴레이 CH1 COM', ''),
 ('펌프1 (+) 연결', '릴레이 CH1 NO', '펌프1 + (빨강)', '접점 닫히면 펌프 ON'),
 ('12V GND 귀환',  '12V 어댑터 −', '펌프1 − (검정)', ''),
 ('플라이백 다이오드', '1N4007 띠(K)', '펌프1 (+) 쪽', '거꾸로면 단락!'),
 ('플라이백 다이오드', '1N4007 반대(A)', '펌프1 (−) 쪽', ''),
]
r = data_rows(ws, r, rows); r += 1
header_row(ws, r, ['단자대 구멍', '같이 꽂는 선들', '', ''], FILL_HDR_C); r += 1
rows = [
 ('(+) 단자', '릴레이 NO1  +  펌프1 빨강선  +  다이오드 띠 쪽 다리', '', ''),
 ('(−) 단자', '12V 어댑터(−)  +  펌프1 검정선  +  다이오드 반대쪽 다리', '', ''),
]
r = data_rows(ws, r, rows); r += 1

section_row(ws, r, '④ 펌프2 — 순환 (12V DC · 릴레이 CH2) ★ 다이오드 포함', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['항목', '출발', '도착', '비고'], FILL_HDR_B); r += 1
rows = [
 ('12V 활선 입력', '12V 어댑터 +', '릴레이 CH2 COM', ''),
 ('펌프2 (+) 연결', '릴레이 CH2 NO', '펌프2 + (빨강)', '접점 닫히면 펌프 ON'),
 ('12V GND 귀환',  '12V 어댑터 −', '펌프2 − (검정)', ''),
 ('플라이백 다이오드', '1N4007 띠(K)', '펌프2 (+) 쪽', '거꾸로면 단락!'),
 ('플라이백 다이오드', '1N4007 반대(A)', '펌프2 (−) 쪽', ''),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r,
  '* 12V 어댑터 1개로 펌프1·2 같이 쓰면 전류 용량 확인 (둘 합쳐 1.5A 이상 권장).\n'
  '* 따로 어댑터 2개 쓰면 더 안정.',
  4); r += 2

section_row(ws, r, '⑤ 히터 — HE-50W (220V AC · 릴레이 CH3)', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['항목', '출발', '도착', '비고'], FILL_HDR_B); r += 1
rows = [
 ('220V 활선 L',   '220V 콘센트 L', '릴레이 CH3 COM', '⚠ 반드시 코드 뽑고 작업'),
 ('히터 L',        '릴레이 CH3 NO', '히터 L', '접점 닫히면 히터 ON'),
 ('중성선 N',      '220V 콘센트 N', '히터 N', '자르지 않음 (OFF에도 통전)'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r,
  '⚠ 220V 안전: 코드 뽑고 작업 · L/N 단자 절연 · 12V와 220V 배선 물리적 분리 · 누전차단 콘센트 권장',
  4, FILL_DANGER, FONT_WARN)

# ============ 시트 4: 핀 요약 ============
ws = wb.create_sheet('핀_요약')
set_widths(ws, [12, 24, 22, 38])
title_row(ws, 1, 'ESP32 핀 사용 요약 — 한눈에', 4)

r = 3
section_row(ws, r, 'A보드 (센서 + LCD)', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['ESP32 핀', '연결', '종류', '비고'], FILL_HDR_A); r += 1
rows = [
 ('3V3',  '3.3V 레일 공급',                          '전원',       '수온·조도용'),
 ('5V',   '5V 레일 공급',                            '전원',       '탁도·pH·수위·LCD용'),
 ('GND',  'GND 레일 공급 (여러 핀)',                 '전원',       '공통 1점 접지'),
 ('IO4',  'DS18B20 DATA',                            '디지털 OneWire', '4.7kΩ 풀업 필수'),
 ('IO13', 'LCD RS',                                  '디지털 출력', ''),
 ('IO14', 'LCD E',                                   '디지털 출력', ''),
 ('IO16', 'LCD D5',                                  '디지털 출력', ''),
 ('IO17', 'LCD D6',                                  '디지털 출력', ''),
 ('IO18', 'LCD D7',                                  '디지털 출력', ''),
 ('IO27', 'LCD D4',                                  '디지털 출력', ''),
 ('IO32', '탁도 분압 노드',                          'ADC1',       'WiFi와 충돌 없음'),
 ('IO33', '조도 분압 노드',                          'ADC1',       ''),
 ('IO34', 'pH 분압 노드',                            'ADC1 (입력전용)', ''),
 ('IO35', '수위 분압 노드',                          'ADC1 (입력전용)', '디지털로 읽음'),
]
r = data_rows(ws, r, rows); r += 2

section_row(ws, r, 'B보드 (릴레이 제어 + BLE)', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['ESP32 핀', '연결', '종류', '비고'], FILL_HDR_B); r += 1
rows = [
 ('5V',   '릴레이 모듈 VCC', '전원', ''),
 ('GND',  '릴레이 모듈 GND', '전원', '공통'),
 ('IO25', '릴레이 IN1 → 펌프1 (급수)', '디지털 출력', 'LOW=ON'),
 ('IO26', '릴레이 IN2 → 펌프2 (순환)', '디지털 출력', 'LOW=ON'),
 ('IO27', '릴레이 IN3 → 히터 (220V)',  '디지털 출력', 'LOW=ON'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* 나머지 B보드 핀은 미사용. 산소(oxygen)는 핀 없는 로직 전용.', 4)

# ============ 시트 5: 부품 목록 ============
ws = wb.create_sheet('부품_목록')
set_widths(ws, [12, 32, 12, 38])
title_row(ws, 1, '부품 목록 (12V 펌프 2개 버전)', 4)

r = 3
section_row(ws, r, 'A보드 측', 4, FILL_HDR_A); r += 1
header_row(ws, r, ['카테고리', '부품', '수량', '비고'], FILL_HDR_A); r += 1
rows = [
 ('보드',  'ESP32 38핀 DevKit + GPIO 확장보드', 1, ''),
 ('보드',  '브레드보드',                        1, '레일 정리용'),
 ('센서',  'DS18B20 (수온)',                    1, '4.7kΩ 풀업저항 1개 동봉/별도'),
 ('센서',  'GL5528 (조도)',                     1, ''),
 ('센서',  'SEN0189 (탁도)',                    1, ''),
 ('센서',  'PH4502C (pH)',                      1, '교정 필요'),
 ('센서',  'XKC-Y25-V (수위)',                  1, ''),
 ('표시',  'LCD 16×2 (HD44780)',                1, ''),
 ('표시',  '10kΩ 가변저항',                     1, 'LCD 콘트라스트'),
 ('저항',  '4.7kΩ',                             1, 'DS18B20 풀업'),
 ('저항',  '10kΩ',                              5, 'LCD 콘트라스트 제외, 분압용·LDR용'),
 ('저항',  '20kΩ',                              3, '탁도·pH·수위 분압 하단'),
 ('저항',  '220Ω',                              1, 'LCD 백라이트'),
 ('커패시터','10µF 전해',                       2, '5V·3.3V 레일 디커플링'),
 ('커패시터','0.1µF 세라믹(104)',               3, '레일×2 + 조도/탁도용'),
 ('커패시터','1µF',                             1, 'pH 안정용'),
 ('배선',  '점퍼선 (오스-오스/오스-암)',         '적당량', ''),
]
r = data_rows(ws, r, rows); r += 2

section_row(ws, r, 'B보드 측', 4, FILL_HDR_B); r += 1
header_row(ws, r, ['카테고리', '부품', '수량', '비고'], FILL_HDR_B); r += 1
rows = [
 ('보드',     'ESP32 WROOM-32',                   1, ''),
 ('스위칭',   '3채널 릴레이 모듈 (LOW 트리거)',   1, '250VAC 이상 정격'),
 ('커패시터', '10µF 전해 + 0.1µF 세라믹',         '각 1', '릴레이 VCC 디커플링'),
 ('전원',     '12V 어댑터 (1.5A 이상)',           1, '펌프1·2 공용 또는 펌프당 1개'),
 ('펌프',     '12V DC 펌프 (SZH-PWAT-040 등)',    2, '급수·순환'),
 ('다이오드', '1N4007',                           2, '★ 펌프1·2 각 1개 플라이백'),
 ('단자대',   '2P 스크류 단자대 (KF301-2P 등)',   2, '펌프별 1개 (납땜 대체)'),
 ('대안',     'Wago 레버 커넥터',                 '대체가능', '단자대 대신'),
 ('히터',     'HE-50W (220V) + 코드(돼지코)',     1, ''),
 ('전원',     'USB 5V 어댑터 (1.5~2A)',           1, 'ESP32 본체'),
]
r = data_rows(ws, r, rows); r += 1
note_block(ws, r, '* B보드 부품 중 ★ 표시는 안정/안전을 위해 꼭 추가해야 하는 항목.', 4)

# ============ 시트 6: 주의사항 ============
ws = wb.create_sheet('주의사항')
set_widths(ws, [6, 70])
title_row(ws, 1, '자주 헷갈리는 7가지 + 안전', 2)

r = 3
header_row(ws, r, ['#', '내용'], FILL_HDR_D); r += 1
rows = [
 (1, 'GND는 한 점으로 묶기 — 보드별로 1점 접지 (A·B는 무선이라 분리 OK)'),
 (2, '5V 센서를 ESP32 IO에 직결 금지 — 반드시 10k+20k 분압 거쳐서 (탁도·pH·수위)'),
 (3, '수온·조도는 분압 X — 3.3V 레일에서 native 3.3V로 그대로 IO에 연결'),
 (4, 'DS18B20 풀업(4.7kΩ) 빠지면 −127℃ 출력 — DATA↔3.3V 사이에 반드시'),
 (5, '다이오드 띠(K) = 펌프 (+) 쪽 — 거꾸로 꽂으면 평상시에도 단락 → 다이오드 즉사'),
 (6, '릴레이 LOW=ON — 시리얼 보고 반대로 동작하면 코드 RELAY_LOW_TRIG=false'),
 (7, '220V 작업은 반드시 코드(돼지코) 뽑고 — L/N 단자 절연·노출 금지'),
]
r = data_rows(ws, r, rows); r += 2

section_row(ws, r, '디커플링 커패시터 설치 요령', 2, FILL_HDR_D); r += 1
note_block(ws, r,
  '· 10µF 전해 + 0.1µF 세라믹을 병렬로\n'
  '· 릴레이/ESP32 등 부하 근처에 다리 짧게\n'
  '· 전해 극성: 긴 다리 = +, 띠 있는 짧은 다리 = −  (거꾸로 꽂으면 부풀거나 터질 수 있음)',
  2)

r += 2
section_row(ws, r, '플라이백 다이오드 설치 요령 (12V 펌프)', 2, FILL_HDR_D); r += 1
note_block(ws, r,
  '· 1N4007 한 개 (1A · 1000V, 펌프당 1개 필수)\n'
  '· 띠(흰색 K) → 펌프 (+) 쪽\n'
  '· 단자대 (+) 구멍에 [릴레이 NO + 펌프 빨강 + 다이오드 띠다리] 같이 꽂기\n'
  '· 단자대 (−) 구멍에 [12V GND + 펌프 검정 + 다이오드 반대다리] 같이 꽂기\n'
  '· 펌프 소비전류 1A 초과면 1N5408(3A) 또는 SR360(쇼트키)',
  2)

wb.save(OUT)
print('saved:', OUT)
